import math
import sys
from re import search

import openpyxl
#from tkinter.ttk import Label

from PySide6.QtCore import QDate, Qt
from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QApplication, QHeaderView, QCheckBox, QTableWidgetItem, \
    QTableWidget, QFileDialog
from qfluentwidgets import CardWidget, PushButton, SearchLineEdit, ComboBox, FastCalendarPicker, TableWidget, \
    setCustomStyleSheet, InfoBar, MessageBox, StrongBodyLabel, PrimaryPushButton, CalendarPicker

from database.ignition_condition_db import IgnitionConditionDB
from project_management.project_dialog import AddProjectDialog, UpdateProjectDialog
from project_management.ground_Info_dialog import GroundInformationBox
from sub_window.ignition_data_entry_interface import IgnitionDataEntryInterface
from utils.custom_button_style import ADD_BUTTON_STYLE,DELETE_BUTTON_STYLE,BATCH_DELETE_BUTTON_STYLE,UPDATE_BUTTON_STYLE,IMPORT_BUTTON_STYLE,EXPORT_BUTTON_STYLE
from database.item_db import itemDB
from database.ppcu_db import ppcuDB
from database.thruster_db import thrusterDB
from database.feedsystem_db import feedsystemDB
from database.ground_info_db import GroundInfoDB
from utils.custom_function import ExcelStyleTool
from utils.ui_components import create_operation_widget, create_action_widget, CustomHeaderView
from utils.ui import MyFastCalendarPicker


#创建项目管理页面基本布局
class ProjectInterface(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('电推进点火数据管理系统')
        self.setObjectName('projectinterface')
        self.items=[]
        self.current_page = 1  # 当前页码
        self.per_page = 20  # 每页显示数量为20
        self.total_pages=1 #总页码
        self.setup_ui()
        self.load_data()
        self.populate_table()

    def setup_ui(self):
        #给界面设定一个整体布局：垂直布局
        layout=QVBoxLayout(self)
        #顶部按钮组
        card_widget=CardWidget(self)
        buttons_layout=QHBoxLayout(card_widget)


        self.addButton=PushButton('新增',self)
        setCustomStyleSheet(self.addButton,ADD_BUTTON_STYLE,ADD_BUTTON_STYLE)
        self.addButton.clicked.connect(self.add_item)#设置关联事件
        self.comboBox=ComboBox(self)
        self.comboBox.setPlaceholderText('选择工质')
        self.comboBox.addItems(['所有工质','氪气','氙气'])
        self.comboBox.setCurrentIndex(-1)
        self.comboBox.activated.connect(self.on_search_change)
        #self.picker=FastCalendarPicker(self)
        self.picker=MyFastCalendarPicker(self)
        #self.picker.setDate(QDate(2026,4,10))
        self.picker.setResetEnabled(True)
        self.picker.setText('请选择日期')
        self.picker.dateChanged.connect(self.on_search_change)

        self.searchInput=SearchLineEdit(self)
        self.searchInput.setPlaceholderText('搜索项目PPCU编号/推力器编号/贮供编号')
        self.searchInput.setFixedWidth(500)
        self.searchInput.searchSignal.connect(self.on_search_change)
        self.searchInput.clearSignal.connect(self.on_search_change)
        self.batch_export_button=PushButton('批量导出',self)
        setCustomStyleSheet(self.batch_export_button,EXPORT_BUTTON_STYLE,EXPORT_BUTTON_STYLE)
        self.batch_export_button.clicked.connect(self.batch_export_items)
        self.batch_import_button=PushButton('批量导入',self)
        setCustomStyleSheet(self.batch_import_button,IMPORT_BUTTON_STYLE,IMPORT_BUTTON_STYLE)
        self.batch_import_button.clicked.connect(self.batch_import_items)
        self.batchDeleteButton=PushButton('批量删除',self)
        setCustomStyleSheet(self.batchDeleteButton,DELETE_BUTTON_STYLE,DELETE_BUTTON_STYLE)
        self.batchDeleteButton.clicked.connect(self.batch_delete)

        buttons_layout.addWidget(self.addButton)
        buttons_layout.addWidget(self.comboBox)
        buttons_layout.addWidget(self.picker)
        buttons_layout.addWidget(self.searchInput)
        buttons_layout.addStretch(1)
        buttons_layout.addWidget(self.batch_import_button)
        buttons_layout.addWidget(self.batch_export_button)
        buttons_layout.addWidget(self.batchDeleteButton)
        layout.addWidget(card_widget)

        #添加table表
        self.table_widget=TableWidget(self)
        #self.table_widget.verticalHeader().setVisible(False)
        self.table_widget.setEditTriggers(QTableWidget.NoEditTriggers)#关闭表格所有编辑触发行为（双击、选中、输入全部失效）
        self.table_widget.setBorderRadius(8)#设置圆角
        self.table_widget.setBorderVisible(True)#设置表格边框可见
        self.table_widget.setViewportMargins(0, 0, 20, 0)
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)  # 设置表头填充满


        #设置表头
        self.table_widget.setColumnCount(12)#设置表格列数
        self.table_widget.setHorizontalHeaderLabels(['','操作项','项目ID','项目名称','PPCU','推力器','贮供','工质','软件版本号','点火日期','点火地点','数据操作'])
        #自定义表头，写入到table_widget中
        self.custom_header=CustomHeaderView(Qt.Orientation.Horizontal,self.table_widget)
        self.table_widget.setHorizontalHeader(self.custom_header)
        layout.addWidget(self.table_widget)

        #添加翻页布局
        self.turn_page_widget=QWidget(self)
        turn_page_layout=QHBoxLayout(self.turn_page_widget)
        self.btn_up_page=PrimaryPushButton("上一页",self)
        self.btn_up_page.clicked.connect(self.up_page)
        self.lab_page=StrongBodyLabel('...')
        self.btn_next_page=PrimaryPushButton("下一页",self)
        self.btn_next_page.clicked.connect(self.next_page)
        turn_page_layout.addStretch(1)
        turn_page_layout.addWidget(self.btn_up_page)
        turn_page_layout.addWidget(self.lab_page)
        turn_page_layout.addWidget(self.btn_next_page)


        layout.addWidget(self.turn_page_widget)#将翻页布局加入垂直布局

        self.setStyleSheet("ProjectInterface{background:white}")
        self.resize(1280,760)






    #设置需填入表格的信息
    def load_data(self,search_text=None,combo_text=None,date_str=None):
        '''从数据库中获取数据'''
        # 操作ignition_item表数据库
        with itemDB() as db:
            if search_text or combo_text or date_str:
                self.items = db.search_items(search_text,combo_text,date_str,self.current_page,self.per_page)
                total_items=db.fetch_filter_items(search_text,combo_text,date_str)
            else:
                #self.items = db.fetch_items()
                self.items=db.fetch_trun_page(self.current_page,self.per_page)
                #总页码数
                total_items=db.fetch_total_items()#总条目数
            self.total_pages=math.ceil(total_items/self.per_page)
        #展示分页效果
        self.display_pagination_effect()



    #展示表格数据
    def populate_table(self):
        '''将数据展示到表格中'''
        self.table_widget.setRowCount(len(self.items))
        for row,item_info in enumerate(self.items):
            self.setup_table_row(row,item_info)

    #遍历需填入表格的信息
    def setup_table_row(self,row,item_info):
        checkbox=QCheckBox()
        checkbox.setStyleSheet('margin:10px')
        self.table_widget.setCellWidget(row,0,checkbox)
        #赋值其他列
        for col,key in enumerate(['project_id','project_name','ppcu_number','thruster_number','feedsystem_number','working_fluid','sw_version','ignition_date','ignition_location','action']):
            value=item_info.get(key,'')
            #单元格赋值
            item=QTableWidgetItem(str(value))# 创建一个单元格，里面放文字
            self.table_widget.setItem(row,col+2,item)

        #在【操作项】列添加ToolButton【编辑】、【导入】、【导出】、【删除】按钮，并创建连接
        operation_widget=create_operation_widget(
            edit_callback=lambda :self.edit_project(item_info['project_id']),
            import_callbak=lambda :self.import_item(item_info['project_id']),
            export_callbak=lambda :self.export_item(item_info['project_id']),
            delete_callback=lambda:self.delete_project(item_info['project_id'])
        )
        self.table_widget.setCellWidget(row,1,operation_widget)

        #在【数据操作】列添加【条件】、【记录】按钮
        action_widget=create_action_widget(
            condition_callback=lambda :self.condition_set(item_info['project_id']),
            record_callback=lambda :self.record_data(item_info['project_id'])
        )
        self.table_widget.setCellWidget(row,11,action_widget)

    def add_item(self):
        '''添加项目信息'''
        dialog=AddProjectDialog(self)
        if dialog.exec():
            #获取【新增项目】对话框中的信息
            project_info=dialog.get_project_info()
            #将获取的信息写入数据库
            try:
                with itemDB() as db:
                    db.add_item(project_info)
                self.filter_conditions()
                self.load_data(self.search_text,self.combo_text,self.date_str)
                self.populate_table()
                InfoBar.success(title='数据添加成功',content=f"{project_info['project_name']}添加成功",duration=3000,parent=self)
            except Exception as e:
                InfoBar.error(title='添加失败',content=str(e),duration=3000,parent=self)

    def edit_project(self,project_id):
        '''编辑项目'''
        dialog=UpdateProjectDialog(project_id,self)
        with itemDB() as db:
            project_info=db.fetch_existing_items(project_id)
        dialog.set_project_info(project_info)
        if dialog.exec():
            #获取【编辑项目】对话框中的信息
            new_project_info=dialog.get_project_info()
            #写入到数据库
            try:
                with itemDB() as db:
                    db.edit_item(new_project_info,project_id)
                self.filter_conditions()
                self.load_data(self.search_text,self.combo_text,self.date_str)
                self.populate_table()
                InfoBar.success(title='编辑成功',content=f"{project_info['project_name']}项目编辑成功",duration=3000,parent=self)
            except Exception as e:
                InfoBar.error(title='编辑失败',content=str(e),duration=3000,parent=self)

    def delete_project(self,project_id):
        '''删除项目信息'''
        #获取project_id对应的项目名称
        with itemDB() as db:
            project_info=db.fetch_existing_items(project_id)
        project_name=project_info['project_name']
        #弹出删除对话框并确认要删除后，通过数据库查询并删除数据
        messagebox=MessageBox(title='确认删除',content=f"是否确定删除{project_name}项目数据？",parent=self)
        if messagebox.exec():
            try:
                with itemDB() as db:
                    db.delete_item(project_id)
                self.load_data()
                self.populate_table()
                InfoBar.success(title='删除成功',content=f'成功删除{project_name}项目数据',duration=3000,parent=self)
            except Exception as e:
                InfoBar.error(title='删除失败',content=str(e),duration=3000,parent=self)

    def batch_delete(self):
        '''批量删除项目信息'''
        self.chk_project_id()
        if len(self.id_collect)==0:
            InfoBar.error(title='未勾选', content='请勾选待删除项', duration=3000, parent=self)
        else:
            messagebox = MessageBox(title='删除确认', content='是否要删除项目信息', parent=self)
            if messagebox.exec():
                try:
                    with itemDB() as db:
                        db.batch_delete_item(self.id_collect)
                    self.load_data()
                    self.populate_table()
                    InfoBar.success(title='删除成功', content=f'成功删除{len(self.id_collect)}项数据', duration=3000, parent=self)
                except Exception as e:
                    InfoBar.error(title='删除失败', content=str(e), duration=3000, parent=self)

    # ========== 导入项目信息 ==========
    def import_item(self,current_id):
        file_path,_=QFileDialog.getOpenFileName(self,'导入项目','','Excel 文件(*.xlsx *.xls)')
        if not file_path:
            InfoBar.error(title='导入取消', content='用户取消了导入操作', duration=3000, parent=self)
            return 
        #加载excel
        skipped_count = 0  # 跳过的数量
        item_excel=openpyxl.load_workbook(file_path)
        #对sheet1进行数据录入
        item_sheet=item_excel.active
        refer_item_header=['项目ID', '项目名称', 'PPCU', '推力器', '贮供', '工质', '软件版本号', '点火日期', '点火地点']
        item_header=[cell.value for cell in item_sheet[1]]
        if item_header!=refer_item_header:
            return
        try:
            with itemDB() as item_db:
                for row in item_sheet.iter_rows(min_row=2,max_row=2,values_only=True):
                    project_id, project_name, ppcu_number, thruster_number, feedsystem_number, working_fluid, sw_version, ignition_date, ignition_location = row
                    if not all([project_id, project_name, working_fluid, ignition_date, ignition_location]):
                        skipped_count+=1
                        continue
                    # 获取ppcu ID
                    with ppcuDB() as ppcu_db:
                        ppcu_id = ppcu_db.fetch_ppcuID(ppcu_number)
                    if ppcu_id is None:
                        skipped_count += 1
                        continue
                    # 获取推力器ID
                    with thrusterDB() as thruster_db:
                        thruster_id = thruster_db.fetch_thrusterID(thruster_number)
                    if thruster_id is None:
                        skipped_count += 1
                        continue
                    # 获取贮供ID
                    with feedsystemDB() as feeds_db:
                        feedsystem_id = feeds_db.fetch_feedsystemID(feedsystem_number)
                    if feedsystem_id is None:
                        skipped_count += 1
                        continue
                    item_info = {
                        'project_id': project_id,
                        'project_name': project_name,
                        'ppcu_id': ppcu_id,
                        'thruster_id': thruster_id,
                        'feedsystem_id': feedsystem_id,
                        'working_fluid': working_fluid,
                        'sw_version': sw_version,
                        'ignition_date': ignition_date,
                        'ignition_location': ignition_location
                    }
                if skipped_count==0:
                    # 检查要导入的数据的project_id是否和当前选择的id一致
                    if current_id != project_id:
                        InfoBar.error(title='导入失败', content='导入的项目ID与当前项目ID不一致', duration=3000,
                                      parent=self)
                    else:
                        item_db.edit_excel_item(item_info, project_id)
                        InfoBar.success(title='导入成功', content=f'成功导入ID编号为{project_id}的数据',
                                        duration=3000, parent=self)
                else:
                    InfoBar.error(title='导入失败',content=f'文件中存在未填项或不匹配项',duration=3000,parent=self)

            self.filter_conditions()
            self.load_data(self.search_text, self.combo_text, self.date_str)
            self.populate_table()
        except Exception as e:
            InfoBar.error(title='导入失败',content=str(e),duration=3000,parent=self)


    # ========== 批量导入项目信息 ==========
    def batch_import_items(self):
        file_path, _ = QFileDialog.getOpenFileName(self, '导入项目', '', 'Excel 文件(*.xlsx *.xls)')
        if not file_path:
            InfoBar.error(title='导入取消', content='用户取消了导入操作', duration=3000, parent=self)
            return
        # 加载excel
        items_excel = openpyxl.load_workbook(filename=file_path)
        items_sheet = items_excel.active  # 获取sheet
        original_header = [cell.value for cell in items_sheet[1]]  # 获取excel中表头数据
        expected_header = ['项目ID', '项目名称', 'PPCU', '推力器', '贮供', '工质', '软件版本号', '点火日期', '点火地点']
        if original_header != expected_header:
            return
        w = MessageBox(title='更新确认', content='是否更新已经存在的信息，点“是”更新，点“否”跳过', parent=self)
        w.yesButton.setText('是')
        w.cancelButton.setText('否')
        if w.exec():
            update_existing = True
        else:
            update_existing = False

        imported_count = 0  # 导入的数量
        skipped_count = 0  # 跳过的数量
        update_count = 0  # 更新的数量
        try:
            with itemDB() as item_db:
                for row in items_sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        break
                    project_id, project_name, ppcu_number, thruster_number, feedsystem_number, working_fluid, sw_version, ignition_date, ignition_location = row
                    # 检查必填字段
                    if not all(
                            [project_id, project_name, working_fluid,ignition_date, ignition_location]):
                        skipped_count += 1
                        continue
                    # 获取ppcu ID
                    with ppcuDB() as ppcu_db:
                        ppcu_id = ppcu_db.fetch_ppcuID(ppcu_number)
                    if ppcu_id is None:
                        skipped_count += 1
                        continue
                    # 获取推力器ID
                    with thrusterDB() as thruster_db:
                        thruster_id = thruster_db.fetch_thrusterID(thruster_number)
                    if thruster_id is None:
                        skipped_count += 1
                        continue
                    # 获取贮供ID
                    with feedsystemDB() as feeds_db:
                        feedsystem_id = feeds_db.fetch_feedsystemID(feedsystem_number)
                    if feedsystem_id is None:
                        skipped_count += 1
                        continue

                    item_info = {
                        'project_id': project_id,
                        'project_name': project_name,
                        'ppcu_id': ppcu_id,
                        'thruster_id': thruster_id,
                        'feedsystem_id': feedsystem_id,
                        'working_fluid': working_fluid,
                        'sw_version': sw_version,
                        'ignition_date': ignition_date,
                        'ignition_location': ignition_location
                    }
                    # 从数据库中查看有没有与导入文件中project_id重复的项
                    existing_id = item_db.fetch_existing_items(project_id)
                    if existing_id:
                        if update_existing:
                            item_db.edit_excel_item(project_info=item_info, project_id=item_info['project_id'])
                            update_count += 1
                        else:
                            skipped_count += 1
                    else:
                        item_db.add_excel_item(item_info)
                        imported_count += 1
            self.filter_conditions()
            self.load_data(self.search_text, self.combo_text, self.date_str)
            self.populate_table()
            InfoBar.success(title='数据导入成功',
                            content=f'成功导入{imported_count}条数据，更新{update_count}条数据，跳过{skipped_count}条数据',
                            duration=3000, parent=self)
        except Exception as e:
            InfoBar.error(title='导入失败', content=str(e), duration=3000, parent=self)

    # ========== 导出项目信息 ==========
    def export_item(self,project_id):
        try:
            excel_item = openpyxl.Workbook()  # 创建一个excel
            #操作sheet1，将项目信息填入sheet1
            with itemDB() as db:
                item=db.fetch_existing_items(project_id)
            sheet_item=excel_item.active
            sheet_item.title='项目信息'
            header=[ '项目名称', 'PPCU', '推力器', '贮供', '工质', '软件版本号', '点火日期', '点火地点']
            sheet_item.append(header)
            item_data=[
                #item['project_id'],
                item['project_name'],
                item['ppcu_number'],
                item['thruster_number'],
                item['feedsystem_number'],
                item['working_fluid'],
                item['sw_version'],
                item['ignition_date'],
                item['ignition_location']
            ]
            sheet_item.append(item_data)

            #操作sheet2，将接地信息填入sheet2
            with GroundInfoDB() as ground_db:
                ground_info=ground_db.fetch_projectid_to_ground(project_id)
            sheet_ground=excel_item.create_sheet(title='接地信息')
            ground_header=['大仓-小仓','PPCU-冷板','PPCU-推力器','PPCU-贮供','母线负-冷板','OC负-冷板','PPCU-大仓MΩ','PPCU-小仓MΩ','通讯地-冷板','备注']
            sheet_ground.append(ground_header)
            ground_data=[
                #ground_info['ground_id'],
                #ground_info['project_id'],
                ground_info['mc_sc'],
                ground_info['ppcu_coldplate'],
                ground_info['ppcu_thruster'],
                ground_info['ppcu_feedsystem'],
                ground_info['busneg_feedsystem'],
                ground_info['oc_feedsystem'],
                ground_info['ppcu_mc'],
                ground_info['ppcu_sc'],
                ground_info['commgnd_feedsystem'],
                ground_info['grounding_notes']
            ]
            sheet_ground.append(ground_data)

            #操作sheet3，将点火条件和点火记录填入sheet3
            with IgnitionConditionDB() as condb:
                con_dict=condb.get_condition_and_record(project_id)
            sheet_condition=excel_item.create_sheet(title='点火条件和结果记录')
            con_header=['点火开始时间', '点火模式', '母线电压', '母线电流', '阳极暂态电压', '阳极稳态电压', '贮供流量控制目标值',
             '开环初期阶段流量校正系数', '加热电流', '触持后加热电流', '触持电流', '励磁电流', '阳极电流上限值',
             '系统总功率上限', 'PID比例常数', '推力器固定启动时长', '点火结束除气时间', '阳极点火时长',
             '电磁阀驱动主/备', '点火条件备注', '示波器1图号', '示波器2图号', '流量调节时间', '92阶段流量', '加热时长',
             '加阳极-阳极浪涌电流', '加阳极-阳极电压最大值', '加阳极-触持电流最小值', '加阳极-母线电流浪涌',
             '阳极暂态电压', '加励磁-阳极浪涌电流', '加励磁-阳极电压最大值', '加励磁-阳极电压下探值',
             '加励磁-母线电流浪涌', '加励磁-母线电流下探', '稳态-阳极稳态电压', '稳态-阳极稳态电流', '稳态-母线电流',
             '稳态-阳极电流峰峰值', '稳态-母线电流峰峰值', '稳态-母线电压峰峰值', '稳态-触持电流峰峰值', '推力',
             '稳态阶段流量', '试验结论', '结论备注']
            sheet_condition.append(con_header)
            for con_dict in con_dict:
                con_data = [
                    con_dict['ignition_time'],
                    con_dict['ignition_mode'],
                    con_dict['busbar_voltage'],
                    con_dict['busbar_current'],
                    con_dict['transient_voltage'],
                    con_dict['steady_voltage'],
                    con_dict['fc_tgt'],
                    con_dict['ol_ifc'],
                    con_dict['heating_current'],
                    con_dict['heating_current2'],
                    con_dict['keep_alive_current'],
                    con_dict['excitation_current'],
                    con_dict['anode_current_limit'],
                    con_dict['system_power_limit'],
                    con_dict['pid_proportional_constant'],
                    con_dict['thruster_duration'],
                    con_dict['pi_cdt'],
                    con_dict['anode_duration'],
                    con_dict['sv_mbs'],
                    con_dict['condition_notes'],
                    con_dict['oscilloscope1_drawnumber'],
                    con_dict['oscilloscope2_drawnumber'],
                    con_dict['flow_adjustment_time'],
                    con_dict['heat_traffic'],
                    con_dict['heat_duration'],
                    con_dict['p1_anode_surge_current'],
                    con_dict['p1_maximum_anode_voltage'],
                    con_dict['p1_minimum_holding_current'],
                    con_dict['p1_bus_current_surge'],
                    con_dict['p2_anode_transient_voltage'],
                    con_dict['p3_anode_surge_current'],
                    con_dict['p3_maximum_anode_voltage'],
                    con_dict['p3_minimum_anode_voltage'],
                    con_dict['p3_bus_current_surge'],
                    con_dict['p3_minimum_bus_current'],
                    con_dict['p4_transient_voltage'],
                    con_dict['p4_anode_steady_current'],
                    con_dict['p4_bus_current'],
                    con_dict['p4_pp_anode_current'],
                    con_dict['p4_pp_bus_current'],
                    con_dict['p4_pp_busbar_voltage'],
                    con_dict['p4_pp_keep_alive_current'],
                    con_dict['thrust'],
                    con_dict['p4_traffic'],
                    con_dict['test_conclusion'],
                    con_dict['record_notes']
                ]
                sheet_condition.append(con_data)



            # 设置excel中数据全局居中和自适应列宽
            ExcelStyleTool.auto_adjust_all_sheets(excel_item)


            file_path,_=QFileDialog.getSaveFileName(self,'保存文件','','Excel 文件(*.xlsx)')
            if file_path:
                excel_item.save(file_path)
                InfoBar.success(title='导出成功',content=f'成功导出{item["project_name"]}项目',duration=3000,parent=self)
            else:
                InfoBar.error(title='导出取消',content='用户取消了文件导出操作',duration=3000,parent=self)

        except Exception as e:
            InfoBar.error(title='导出失败',content=str(e),duration=3000,parent=self)

    # ========== 批量导出项目信息 ==========
    def batch_export_items(self):
        w=MessageBox(title='导出确认',content='是否批量导出项目信息',parent=self)
        if w.exec():
            try:
                self.filter_conditions()
                self.chk_project_id()
                #从数据库中获取所有项目信息
                with itemDB() as db:
                    items=db.batch_export_items(self.search_text,self.combo_text,self.date_str,self.id_collect)
                #将项目信息写入到excel
                item_excel=openpyxl.Workbook() #创建一个excel
                item_sheet=item_excel.active
                item_sheet.title='项目信息'
                item_header = ['项目ID', '项目名称', 'PPCU', '推力器', '贮供', '工质', '软件版本号', '点火日期', '点火地点'] #表头信息
                item_sheet.append(item_header) #将表头信息填入excel
                for item in items:
                    per_row_item=[
                        item['project_id'],
                        item['project_name'],
                        item['ppcu_number'],
                        item['thruster_number'],
                        item['feedsystem_number'],
                        item['working_fluid'],
                        item['sw_version'],
                        item['ignition_date'],
                        item['ignition_location']
                    ]
                    item_sheet.append(per_row_item)
                #保存excel
                file_path,_=QFileDialog.getSaveFileName(self,'保存文件','项目文件','Excel 文件 (*.xlsx)') #设置文件保存的路径
                if file_path:
                    item_excel.save(file_path)
                    InfoBar.success(title='导出成功',content='批量导出成功',duration=2000,parent=self)
                else:
                    InfoBar.warning(title='导出取消',content='用户取消了导出',duration=2000,parent=self)
            except Exception as e:
                InfoBar.error(title='导出失败',content=str(e),duration=2000,parent=self)




    def condition_set(self,project_id):
        ground_info_box=GroundInformationBox(project_id,self)
        if ground_info_box.exec():
            # 获取各个输入框中的文本
            ground_info = ground_info_box.get_grounding_info()
            #将文本保存进数据库,去判断project_id,如果已经存在，则对数据进行更新；否则的话添加
            with GroundInfoDB() as db:
                if db.fetchbool_projectid_to_ground(project_id):
                    db.update_ground_info(ground_info)
                    InfoBar.success(title='更新成功',content=f'已成功更新项目ID为{ground_info["project_id"]}的接地信息',duration=3000,parent=self)
                else:
                    db.add_ground_info(ground_info)
                    InfoBar.success(title='添加成功',content=f'已成功添加项目ID为{ground_info["project_id"]}的接地信息',duration=3000,parent=self)
            #将数据库中该条记录展示在dialog中
        else:
            print('用户点击了取消')
    def record_data(self,project_id):
        self.ignition_data_window=IgnitionDataEntryInterface(project_id)
        self.ignition_data_window.show()



    def chk_project_id(self):
        '''获取被勾选的project_id，并存放在列表中'''
        # 创建一个空表,接收勾选的id号
        self.id_collect = []
        # 获取列表的行数
        col = self.table_widget.rowCount()
        for row in range(col):
            checkbox = self.table_widget.cellWidget(row, 0)
            if not checkbox:
                return
            if checkbox.isChecked():
                item = self.table_widget.item(row, 2).text()
                if item is None:
                    return
                project_id = int(item)
                self.id_collect.append(project_id)  # 将所有勾选的项的project_id都集中在一起

    def up_page(self):
        '''翻至上一页'''
        if self.current_page>1:
            self.current_page -= 1
            self.on_page_change()

    def next_page(self):
        '''翻至下一页'''
        if self.current_page<self.total_pages:
            self.current_page+=1
            self.on_page_change()

    #设置【上一页】、【下一页】、页码信息显示的效果
    def display_pagination_effect(self):
        if self.total_pages:
            self.btn_up_page.setEnabled(self.current_page>1)
            self.btn_next_page.setEnabled(self.current_page<self.total_pages)
            self.lab_page.setText(f"第{self.current_page}页，总共{self.total_pages}页")
        else:
            self.lab_page.setText('总共0页')
            self.btn_up_page.setDisabled(True)
            self.btn_next_page.setDisabled(True)

    def on_page_change(self):
        '''date=self.picker.date
        self.date_str=date.toString('yyyy-MM-dd') if date.isValid() else ''
        self.search_text = self.searchInput.text().strip()
        self.combo_text = self.comboBox.currentText()'''
        self.filter_conditions()
        self.load_data(self.search_text,self.combo_text,self.date_str)
        self.populate_table()

    def on_search_change(self):
        self.current_page=1
        self.on_page_change()

    def filter_conditions(self):
        '''获取筛选条件'''
        date = self.picker.date
        self.date_str = date.toString('yyyy-MM-dd') if date.isValid() else ''
        # print(date_str)
        self.search_text = self.searchInput.text().strip()
        self.combo_text = self.comboBox.currentText()

if __name__=="__main__":
    app=QApplication(sys.argv)
    window=ProjectInterface()
    window.show()
    #window.showMaximized()
    sys.exit(app.exec())
