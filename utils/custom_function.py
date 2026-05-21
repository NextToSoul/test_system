import re

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def is_double(txt):
    '''
    对浮点数进行异常捕获判断
    输入框中可以为空，也可以填浮点数，其余格式均非法
    '''
    if txt is None:
        return True
    s=str(txt).strip()
    if not s:
        return True

    try:
        float(s)
        return True
    except ValueError:
        return False

def to_none(s):
    '''如果为空字符串转None；否则返回原字符串'''
    return s if s!='' else None

# ========自定义Excel风格工具========
class ExcelStyleTool:

    @staticmethod
    def auto_adjust_and_center(ws):
        '''
        1. 所有列根据内容自适应宽度（中文适配）
        2. 所有单元格水平+垂直居中
        :param ws: 传入的sheet工作表
        :return:
        '''
        #全局居中
        alignment=Alignment(horizontal='center',vertical='center')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment=alignment

        #自适应列宽
        for col in ws.columns:
            max_width=0
            col_letter=get_column_letter(col[0].column)
            for cell in col:
                if not cell.value:
                    continue
                s=str(cell.value)
                #匹配中文字符
                cn_chars=re.findall(r'[\u4e00-\u9fa5]',s)
                cn_len=len(cn_chars)
                en_len=len(s)-cn_len
                #中文按2个字符宽度计算
                curr_width=en_len+cn_len*2
                if curr_width>max_width:
                    max_width=curr_width
                #预留两个字符边距
                ws.column_dimensions[col_letter].width=max_width+2

    @staticmethod
    def auto_adjust_all_sheets(wb):
        '''
        将效果批量应用于excel中的所有sheet表
        :param wb: 传入的excel文件
        :return:
        '''
        for ws in wb.worksheets:
            ExcelStyleTool.auto_adjust_and_center(ws)
